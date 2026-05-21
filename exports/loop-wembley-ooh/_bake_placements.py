"""Emit data/placements.json from the V5 workbook.

Joins BOOKED - LU 48s Site List (lat/lng + line + platform direction) with
RECOMMENDED CREATIVE PLACEMENT (final A/B allocation) on Frame ID.
"""

import json
import pathlib
from openpyxl import load_workbook

HERE = pathlib.Path(__file__).parent
DATA_DIR = HERE / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

SRC = HERE / "BOOKED_CREATIVE_REC_V5_FINAL_21.05.26.xlsx"

COPY_LINES = {
    "A": "Screaming Encouraged",
    "B": "Hear for a good time and a long time",
}


def main() -> None:
    print(f"Loading {SRC}")
    wb = load_workbook(SRC, data_only=True)

    # First, build a frame_id -> A|B map from REC sheet.
    rec = wb["RECOMMENDED CREATIVE PLACEMENT"]
    frame_to_creative: dict[str, str] = {}
    for r in range(2, rec.max_row + 1):
        creative = rec.cell(row=r, column=1).value
        frame = rec.cell(row=r, column=5).value
        if creative is None or frame is None:
            continue
        frame_to_creative[str(int(frame)) if isinstance(frame, float) else str(frame).strip()] = str(creative).strip()
    print(f"REC sheet: {len(frame_to_creative)} allocations")

    # Now read the booked site list for lat/lng + line/platform.
    booked = wb["BOOKED - LU 48s Site List"]
    headers = [booked.cell(row=1, column=c).value for c in range(1, booked.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}
    needed = ["Location Description", "Town", "Frame/Depot ID", "Postcode", "Latitude", "Longitude", "Station"]
    for n in needed:
        assert n in idx, f"missing column {n} in booked sheet (headers={headers})"

    placements: list[dict] = []
    for r in range(2, booked.max_row + 1):
        row_vals = [booked.cell(row=r, column=c + 1).value for c in range(len(headers))]
        if not any(row_vals):
            continue
        frame_raw = row_vals[idx["Frame/Depot ID"]]
        if frame_raw is None:
            continue
        frame_id = str(int(frame_raw)) if isinstance(frame_raw, float) else str(frame_raw).strip()
        creative = frame_to_creative.get(frame_id)
        if creative is None:
            print(f"  WARN: booked row {r} frame {frame_id} has no allocation; skipping")
            continue
        lat = row_vals[idx["Latitude"]]
        lng = row_vals[idx["Longitude"]]
        if lat is None or lng is None:
            print(f"  WARN: booked row {r} frame {frame_id} missing lat/lng; skipping")
            continue
        station = (row_vals[idx["Station"]] or "").strip()
        desc = (row_vals[idx["Location Description"]] or "").strip()
        town = (row_vals[idx["Town"]] or "").strip()
        postcode = (row_vals[idx["Postcode"]] or "").strip() if row_vals[idx["Postcode"]] else ""

        # Derive line + platform from the description for the HUD.
        line, platform = _parse_description(desc)

        placements.append({
            "frameId": frame_id,
            "station": station,
            "town": town,
            "postcode": postcode,
            "line": line,
            "platform": platform,
            "description": desc,
            "lat": float(lat),
            "lng": float(lng),
            "creative": creative,
            "copy": COPY_LINES[creative],
        })

    placements.sort(key=lambda p: (p["creative"], p["station"], p["frameId"]))
    counts = {"A": 0, "B": 0}
    for p in placements:
        counts[p["creative"]] += 1
    print(f"Placements: {len(placements)}  {counts}")
    assert counts == {"A": 20, "B": 19}, counts

    out = DATA_DIR / "placements.json"
    out.write_text(json.dumps(placements, indent=2), encoding="utf-8")
    size_kb = out.stat().st_size / 1024
    print(f"Wrote {out}  ({size_kb:.1f} KB)")


def _parse_description(desc: str) -> tuple[str, str]:
    """e.g. 'JUBILEE LINE PLATFORM 1 NORTHBOUND' -> ('Jubilee', 'Pf 1 NB')."""
    if not desc:
        return ("", "")
    line_end = desc.find(" LINE")
    if line_end == -1:
        # DLR is described as "DOCKLANDS RAILWAY PLATFORM 1 SOUTHBOUND"
        if "DOCKLANDS" in desc:
            return ("DLR", _parse_platform(desc))
        return ("", desc)
    line_raw = desc[:line_end].strip()
    # "CIRCLE & HAMMERSMITH" stays as-is
    line_name = " ".join(w.capitalize() for w in line_raw.split())
    line_name = line_name.replace(" & ", " & ").replace("And", "&")
    # Tidy known names
    rename = {
        "Northern": "Northern",
        "Jubilee": "Jubilee",
        "Metropolitan": "Metropolitan",
        "Bakerloo": "Bakerloo",
        "Piccadilly": "Piccadilly",
        "Victoria": "Victoria",
        "Central": "Central",
        "Circle & Hammersmith": "Circle & Hammersmith",
    }
    line_name = rename.get(line_name, line_name)
    platform = _parse_platform(desc[line_end:])
    return (line_name, platform)


def _parse_platform(rest: str) -> str:
    rest = rest.upper()
    # find "PLATFORM N <direction>"
    pf = ""
    if "PLATFORM" in rest:
        after = rest.split("PLATFORM", 1)[1].strip()
        parts = after.split()
        if parts:
            pf_num = parts[0]
            direction = ""
            if len(parts) > 1:
                d = parts[1]
                dirmap = {
                    "NORTHBOUND": "NB",
                    "SOUTHBOUND": "SB",
                    "EASTBOUND": "EB",
                    "WESTBOUND": "WB",
                }
                direction = dirmap.get(d, "")
            pf = f"Pf {pf_num} {direction}".strip()
    # Tag Charing X branch on Northern
    if "CHARING X" in rest:
        pf += " (Charing X)"
    return pf


if __name__ == "__main__":
    main()
